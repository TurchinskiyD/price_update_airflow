import os

import openpyxl
import requests

import main_adr
import main_atl
import main_dasmart
import main_dosp
import main_kemp
import main_norf
import main_outfit
import main_shamb
import main_swa
import main_trp


def download_file(url, name):
    file_path = os.path.join("price/", name)

    # виконати запит GET до сервера та отримати відповідь
    try:
        response = requests.get(url)
        response.raise_for_status()

        # зберегти вміст відповіді в файл
        with open(file_path, 'wb') as file:
            file.write(response.content)

        print(f'Файл {name} було успішно завантажено та перезаписано.')

    except requests.exceptions.RequestException as e:
        print(f'Виникла помилка під час завантаження {name}: {str(e)}')


data_for_update = main_outfit.outfit_file_operation()


def add_for_update(name, dictionary):
    dict_temp = dictionary
    counter = 0
    for key, value in dict_temp.items():
        if key in data_for_update:
            try:
                data_for_update[key]['available'] = value['available']
                data_for_update[key]['price'] = value['price']
                counter += 1
            except Exception as e:
                print(f"Помилка при обробці словника {name}: {e}")
                break
    print(f"В каталозі Аутфіттер оновлено {counter} товарів. "
          f"Загльна кількість товарів в прайсі {name} {len(dictionary)} елементів.")


def create_price_xlsx(data_outfitter):
    workbook = openpyxl.Workbook()

    # вибираємо активний лист
    worksheet = workbook.active

    # додаємо заголовки стовпців
    worksheet['A1'] = 'Артикул'
    worksheet['B1'] = 'Наличие'
    worksheet['C1'] = 'Цена'
    worksheet['D1'] = 'Назва'

    keys = list(data_outfitter.keys())

    for i in range(len(keys)):
        key = keys[i]

        row_num = i + 2  # рядок для запису даних, починаємо з другого рядка
        worksheet.cell(row=row_num, column=1, value=key)
        worksheet.cell(row=row_num, column=2, value=data_outfitter[key]['available'])
        worksheet.cell(row=row_num, column=3, value=data_outfitter[key]['price'])
        worksheet.cell(row=row_num, column=4, value=data_outfitter[key]['name'])

    workbook.save("price_update.xlsx")


if __name__ == "__main__":

    link_list = [('https://outfitter.in.ua/content/export/1dc57c3051db9aa40919d7d71ef7b23e.xlsx', 'outfitter.xlsx'),
                 ('https://www.dropbox.com/s/qzg0r695jx3m8iw/rozn.XLS?dl=1', 'shambala.xls'),
                 ('https://tramp.ua/content/export/wholesale/tramp.ua_8f14e45fceea167a5a36dedd4bea2543.xlsx',
                  'tramp.xlsx'),
                 ('https://uabest.com.ua/content/export/36.xml', 'dospehi.xml'),
                 ('https://atlantmarket.com.ua/price1/prom/atlantmarketprom(false).xml', 'atlantmarket.xml'),
                 ('https://opt.adrenalin.od.ua/Adrenalin_stock.xml', 'adrenalin.xml'),
                 ('https://sva.com.ua/content/export/80.xlsx', 'swa.xlsx'),
                 ('https://www.dasmart.com.ua/content/export/e7052aba0549f32c1a825bf48539397b.xml', 'dasmart.xml')]

    for url_link, name_link in link_list:
        download_file(url_link, name_link)

    print(f'Кількість товарів в каталозі Аутфіттер {len(data_for_update)}')

    functions_list = [("Атлант", main_atl.atl_file_operation),
                      ("Кемпінг", main_kemp.kemping_file_operation),
                      ("Шамбала", main_shamb.shamb_file_operation),
                      ("Трамп", main_trp.trp_file_operation),
                      ("Сва", main_swa.swa_file_operation),
                      ("Доспехи", main_dosp.dosp_file_operation),
                      ("Норфін", main_norf.norf_file_operation),
                      ("Адреналін", main_adr.adr_file_operation),
                      ("Дасмарт", main_dasmart.dasmart_file_operation)]

    for name_func, func in functions_list:
        result_dict = func()
        add_for_update(name_func, result_dict)

    create_price_xlsx(data_for_update)



